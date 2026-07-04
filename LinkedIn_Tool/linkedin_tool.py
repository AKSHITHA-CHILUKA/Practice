from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # Optional dependency.
    Workbook = None
    load_workbook = None


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_STORE = BASE_DIR / "prospects.json"
XLSX_HEADERS = ["name", "title", "company", "topics", "profile_url", "status", "note", "last_updated"]
STOPWORDS = {
    "a",
    "about",
    "after",
    "all",
    "and",
    "any",
    "are",
    "as",
    "at",
    "be",
    "been",
    "before",
    "but",
    "by",
    "can",
    "could",
    "for",
    "from",
    "have",
    "has",
    "he",
    "her",
    "his",
    "how",
    "i",
    "if",
    "in",
    "into",
    "is",
    "it",
    "its",
    "just",
    "may",
    "me",
    "more",
    "my",
    "need",
    "new",
    "not",
    "of",
    "on",
    "or",
    "our",
    "out",
    "people",
    "product",
    "profile",
    "s",
    "she",
    "should",
    "so",
    "some",
    "than",
    "that",
    "the",
    "their",
    "them",
    "there",
    "these",
    "they",
    "this",
    "to",
    "up",
    "us",
    "was",
    "we",
    "what",
    "when",
    "where",
    "who",
    "why",
    "will",
    "with",
    "work",
    "would",
    "you",
    "your",
}


@dataclass
class Prospect:
    name: str
    title: str = ""
    company: str = ""
    topics: list[str] = field(default_factory=list)
    profile_url: str = ""
    status: str = "to_contact"
    note: str = ""
    last_updated: str = ""


def now_utc() -> str:
    return datetime.utcnow().isoformat(timespec="seconds") + "Z"


def cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_topics(topics: str | list[str]) -> list[str]:
    values = topics if isinstance(topics, list) else topics.split(",")
    return sorted({value.strip().lower() for value in values if value.strip()})


def load_store(store_path: Path) -> list[Prospect]:
    if not store_path.exists():
        return []

    with store_path.open("r", encoding="utf-8") as handle:
        raw_items = json.load(handle)

    prospects: list[Prospect] = []
    for item in raw_items:
        topics = item.get("topics", [])
        if isinstance(topics, str):
            topics = [value.strip() for value in topics.split(",") if value.strip()]

        prospects.append(
            Prospect(
                name=cell_text(item.get("name")),
                title=cell_text(item.get("title")),
                company=cell_text(item.get("company")),
                topics=normalize_topics(topics),
                profile_url=cell_text(item.get("profile_url")),
                status=cell_text(item.get("status")) or "to_contact",
                note=cell_text(item.get("note")),
                last_updated=cell_text(item.get("last_updated")),
            )
        )

    return prospects


def save_store(store_path: Path, prospects: list[Prospect]) -> None:
    payload = [asdict(prospect) for prospect in prospects]
    with store_path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, ensure_ascii=True)
        handle.write("\n")


def prospect_to_row(prospect: Prospect) -> dict[str, Any]:
    return {
        "name": prospect.name,
        "title": prospect.title,
        "company": prospect.company,
        "topics": ", ".join(prospect.topics),
        "profile_url": prospect.profile_url,
        "status": prospect.status,
        "note": prospect.note,
        "last_updated": prospect.last_updated,
    }


def prospect_from_row(row: dict[str, Any]) -> Prospect | None:
    name = cell_text(row.get("name") or row.get("Name"))
    if not name:
        return None

    topics = row.get("topics") or row.get("interests") or ""
    return Prospect(
        name=name,
        title=cell_text(row.get("title") or row.get("role")),
        company=cell_text(row.get("company")),
        topics=normalize_topics(topics if isinstance(topics, list) else str(topics)),
        profile_url=cell_text(row.get("profile_url") or row.get("url")),
        status=cell_text(row.get("status")) or "to_contact",
        note=cell_text(row.get("note")),
        last_updated=cell_text(row.get("last_updated")) or now_utc(),
    )


def extract_keywords(text: str) -> list[str]:
    words = re.findall(r"[a-zA-Z][a-zA-Z0-9+#.-]{2,}", text.lower())
    return sorted({word for word in words if word not in STOPWORDS and not word.isdigit()})


def prospect_haystack(prospect: Prospect) -> str:
    return " ".join([prospect.name, prospect.title, prospect.company, " ".join(prospect.topics), prospect.note]).lower()


def match_score(prospect: Prospect, keywords: Iterable[str]) -> tuple[int, list[str]]:
    haystack = prospect_haystack(prospect)
    matched: list[str] = []
    score = 0
    topic_set = {topic.lower() for topic in prospect.topics}

    for keyword in keywords:
        if keyword in haystack:
            matched.append(keyword)
            score += 2
            if keyword in topic_set:
                score += 3
        if keyword in prospect.name.lower() or keyword in prospect.title.lower() or keyword in prospect.company.lower():
            score += 2

    if prospect.status == "to_contact":
        score += 1

    return score, matched


def rank_prospects(prospects: list[Prospect], keywords: Iterable[str]) -> list[tuple[int, Prospect, list[str]]]:
    ranked = []
    for prospect in prospects:
        score, matched = match_score(prospect, keywords)
        ranked.append((score, prospect, matched))
    return sorted(ranked, key=lambda item: (item[0], item[1].name.lower()), reverse=True)


def format_prospect(prospect: Prospect) -> str:
    topics = ", ".join(prospect.topics) if prospect.topics else "-"
    note = f" | note: {prospect.note}" if prospect.note else ""
    url = f" | {prospect.profile_url}" if prospect.profile_url else ""
    return (
        f"{prospect.name} | {prospect.title or '-'} | {prospect.company or '-'} | "
        f"topics: {topics} | status: {prospect.status}{note}{url}"
    )


def read_text_input(text: str | None) -> str:
    if text is not None:
        return text
    if sys.stdin.isatty():
        raise SystemExit("No text provided. Pass --text or pipe/paste text via stdin.")
    return sys.stdin.read()


def load_csv_rows(csv_path: Path) -> list[Prospect]:
    prospects: list[Prospect] = []
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            prospect = prospect_from_row(row)
            if prospect is not None:
                prospects.append(prospect)
    return prospects


def load_xlsx_rows(xlsx_path: Path) -> list[Prospect]:
    if load_workbook is None:
        raise SystemExit("openpyxl is required for XLSX support. Install dependencies and try again.")

    workbook = load_workbook(xlsx_path)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [cell_text(value).lower() for value in next(rows, [])]
    prospects: list[Prospect] = []

    for row_values in rows:
        row = {
            headers[index]: row_values[index]
            for index in range(min(len(headers), len(row_values)))
            if headers[index]
        }
        prospect = prospect_from_row(row)
        if prospect is not None:
            prospects.append(prospect)

    return prospects


def export_xlsx_rows(xlsx_path: Path, prospects: list[Prospect]) -> None:
    if Workbook is None:
        raise SystemExit("openpyxl is required for XLSX support. Install dependencies and try again.")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Prospects"
    sheet.append(XLSX_HEADERS)
    for prospect in prospects:
        row = prospect_to_row(prospect)
        sheet.append([row[column] for column in XLSX_HEADERS])
    workbook.save(xlsx_path)


def add_prospect(args: argparse.Namespace) -> None:
    store_path = Path(args.store)
    prospects = load_store(store_path)
    prospect = Prospect(
        name=args.name.strip(),
        title=args.title.strip(),
        company=args.company.strip(),
        topics=normalize_topics(args.topics),
        profile_url=args.profile_url.strip(),
        status=args.status,
        note=args.note.strip(),
        last_updated=now_utc(),
    )
    prospects.append(prospect)
    save_store(store_path, prospects)
    print(f"Saved prospect: {prospect.name}")


def import_csv(args: argparse.Namespace) -> None:
    store_path = Path(args.store)
    prospects = load_store(store_path)
    prospects.extend(load_csv_rows(Path(args.csv)))
    save_store(store_path, prospects)
    print(f"Imported prospects from {args.csv}")


def import_xlsx(args: argparse.Namespace) -> None:
    store_path = Path(args.store)
    prospects = load_store(store_path)
    prospects.extend(load_xlsx_rows(Path(args.xlsx)))
    save_store(store_path, prospects)
    print(f"Imported prospects from {args.xlsx}")


def export_xlsx(args: argparse.Namespace) -> None:
    prospects = load_store(Path(args.store))
    export_xlsx_rows(Path(args.xlsx), prospects)
    print(f"Exported prospects to {args.xlsx}")


def list_prospects(args: argparse.Namespace) -> None:
    prospects = load_store(Path(args.store))
    if not prospects:
        print("No prospects saved yet.")
        return
    for prospect in prospects:
        print(format_prospect(prospect))


def suggest_prospects(args: argparse.Namespace) -> None:
    prospects = load_store(Path(args.store))
    interests = normalize_topics(args.interests)
    if not prospects:
        print("No prospects saved yet.")
        return

    ranked = rank_prospects(prospects, interests)
    for score, prospect, _ in ranked[: args.limit]:
        if score <= 0:
            continue
        print(f"{score:>3} | {format_prospect(prospect)}")


def match_text(args: argparse.Namespace) -> None:
    prospects = load_store(Path(args.store))
    if not prospects:
        print("No prospects saved yet.")
        return

    text = read_text_input(args.text)
    keywords = extract_keywords(text)
    if not keywords:
        print("No meaningful keywords found in the provided text.")
        return

    ranked = rank_prospects(prospects, keywords)
    for score, prospect, matched in ranked[: args.limit]:
        if score <= 0:
            continue
        matched_text = f" | matched: {', '.join(matched[:8])}" if matched else ""
        print(f"{score:>3} | {format_prospect(prospect)}{matched_text}")


def draft_message(args: argparse.Namespace) -> None:
    prospect = Prospect(
        name=args.name,
        title=args.title,
        company=args.company,
        topics=normalize_topics(args.topics),
        profile_url=args.profile_url,
        note=args.note,
    )
    interests = normalize_topics(args.interests)
    topic_phrase = ", ".join(interests[:3]) if interests else "your work"
    shared_topics = ", ".join(sorted(set(prospect.topics).intersection(interests)))

    lines = [
        f"Hi {prospect.name},",
        "",
        f"I came across your profile and was interested in your work{f' at {prospect.company}' if prospect.company else ''}.",
    ]
    if shared_topics:
        lines.append(f"We seem to share interests in {shared_topics}.")
    else:
        lines.append(f"I am currently learning more about {topic_phrase}.")
    lines.extend([
        "I would appreciate connecting and learning from your experience.",
        "",
        "Best regards,",
        args.your_name,
    ])
    print("\n".join(lines))


def mark_status(args: argparse.Namespace) -> None:
    store_path = Path(args.store)
    prospects = load_store(store_path)
    updated = False

    for prospect in prospects:
        if prospect.name.lower() == args.name.lower():
            prospect.status = args.status
            prospect.note = args.note or prospect.note
            prospect.last_updated = now_utc()
            updated = True
            break

    if not updated:
        raise SystemExit(f"No prospect found with name: {args.name}")

    save_store(store_path, prospects)
    print(f"Updated status for {args.name} -> {args.status}")


def init_store(args: argparse.Namespace) -> None:
    store_path = Path(args.store)
    if store_path.exists():
        print(f"Store already exists: {store_path}")
        return
    save_store(store_path, [])
    print(f"Created store: {store_path}")


class ProspectManagerApp(tk.Tk):
    def __init__(self, store_path: Path) -> None:
        super().__init__()
        self.title("LinkedIn Prospect Tool")
        self.geometry("1180x760")
        self.minsize(980, 680)
        self.store_path = store_path
        self.prospects: list[Prospect] = []

        self.style = ttk.Style(self)
        if "vista" in self.style.theme_names():
            self.style.theme_use("vista")

        self.columnconfigure(0, weight=1)
        self.columnconfigure(1, weight=2)
        self.rowconfigure(0, weight=1)

        left_frame = ttk.Frame(self, padding=12)
        right_frame = ttk.Frame(self, padding=12)
        left_frame.grid(row=0, column=0, sticky="nsew")
        right_frame.grid(row=0, column=1, sticky="nsew")

        left_frame.columnconfigure(0, weight=1)
        left_frame.rowconfigure(1, weight=1)
        right_frame.columnconfigure(0, weight=1)
        right_frame.rowconfigure(5, weight=1)
        right_frame.rowconfigure(6, weight=1)

        ttk.Label(left_frame, text="Prospects", font=("Segoe UI", 18, "bold")).grid(row=0, column=0, sticky="w", pady=(0, 8))
        list_container = ttk.Frame(left_frame)
        list_container.grid(row=1, column=0, sticky="nsew")
        list_container.columnconfigure(0, weight=1)
        list_container.rowconfigure(0, weight=1)

        self.listbox = tk.Listbox(list_container, height=24, activestyle="dotbox")
        scrollbar = ttk.Scrollbar(list_container, orient="vertical", command=self.listbox.yview)
        self.listbox.configure(yscrollcommand=scrollbar.set)
        self.listbox.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.listbox.bind("<<ListboxSelect>>", self.on_select)

        self.summary_var = tk.StringVar(value="No prospects loaded.")
        ttk.Label(left_frame, textvariable=self.summary_var, wraplength=320).grid(row=2, column=0, sticky="w", pady=(8, 0))

        ttk.Label(right_frame, text="Prospect Details", font=("Segoe UI", 18, "bold")).grid(row=0, column=0, sticky="w", pady=(0, 8))

        fields = ttk.Frame(right_frame)
        fields.grid(row=1, column=0, sticky="ew")
        fields.columnconfigure(0, weight=1)
        fields.columnconfigure(1, weight=1)

        self.name_var = tk.StringVar()
        self.title_var = tk.StringVar()
        self.company_var = tk.StringVar()
        self.topics_var = tk.StringVar()
        self.profile_url_var = tk.StringVar()
        self.status_var = tk.StringVar(value="to_contact")
        self.note_var = tk.StringVar()

        self.add_field(fields, "Name", self.name_var, 0, 0)
        self.add_field(fields, "Title", self.title_var, 0, 1)
        self.add_field(fields, "Company", self.company_var, 1, 0)
        self.add_field(fields, "Topics", self.topics_var, 1, 1)
        self.add_field(fields, "Profile URL", self.profile_url_var, 2, 0, colspan=2)
        self.add_field(fields, "Status", self.status_var, 3, 0)
        self.add_field(fields, "Note", self.note_var, 3, 1)

        button_bar = ttk.Frame(right_frame)
        button_bar.grid(row=2, column=0, sticky="ew", pady=(12, 8))
        for index in range(4):
            button_bar.columnconfigure(index, weight=1)

        ttk.Button(button_bar, text="Add / Update", command=self.save_current).grid(row=0, column=0, sticky="ew", padx=(0, 6))
        ttk.Button(button_bar, text="Delete", command=self.delete_selected).grid(row=0, column=1, sticky="ew", padx=6)
        ttk.Button(button_bar, text="Refresh", command=self.reload).grid(row=0, column=2, sticky="ew", padx=6)
        ttk.Button(button_bar, text="Clear", command=self.clear_form).grid(row=0, column=3, sticky="ew", padx=(6, 0))

        import_bar = ttk.Frame(right_frame)
        import_bar.grid(row=3, column=0, sticky="ew", pady=(0, 8))
        for index in range(4):
            import_bar.columnconfigure(index, weight=1)

        ttk.Button(import_bar, text="Import CSV", command=self.import_csv_dialog).grid(row=0, column=0, sticky="ew", padx=(0, 6))
        ttk.Button(import_bar, text="Import XLSX", command=self.import_xlsx_dialog).grid(row=0, column=1, sticky="ew", padx=6)
        ttk.Button(import_bar, text="Export XLSX", command=self.export_xlsx_dialog).grid(row=0, column=2, sticky="ew", padx=6)
        ttk.Button(import_bar, text="Match Text", command=self.match_text_dialog).grid(row=0, column=3, sticky="ew", padx=(6, 0))

        ttk.Label(right_frame, text="Paste profile text and click Match Text to rank prospects.").grid(row=4, column=0, sticky="w", pady=(0, 6))
        self.profile_text = tk.Text(right_frame, height=12, wrap="word")
        self.profile_text.grid(row=5, column=0, sticky="nsew")

        self.result_box = tk.Text(right_frame, height=12, wrap="word", state="disabled")
        self.result_box.grid(row=6, column=0, sticky="nsew", pady=(8, 0))

        self.reload()

    def add_field(self, parent: ttk.Frame, label: str, variable: tk.StringVar, row: int, column: int, colspan: int = 1) -> None:
        wrapper = ttk.Frame(parent)
        wrapper.grid(row=row, column=column, sticky="ew", padx=(0, 8), pady=(0, 8), columnspan=colspan)
        wrapper.columnconfigure(0, weight=1)
        ttk.Label(wrapper, text=label).grid(row=0, column=0, sticky="w")
        ttk.Entry(wrapper, textvariable=variable).grid(row=1, column=0, sticky="ew")

    def reload(self) -> None:
        self.prospects = load_store(self.store_path)
        self.refresh_list()
        self.summary_var.set(f"{len(self.prospects)} prospects stored in {self.store_path.name}.")

    def refresh_list(self) -> None:
        self.listbox.delete(0, tk.END)
        for prospect in self.prospects:
            self.listbox.insert(tk.END, f"{prospect.name} | {prospect.title or '-'} | {prospect.status}")

    def selected_prospect(self) -> Prospect | None:
        selected = self.listbox.curselection()
        if not selected:
            return None
        index = int(selected[0])
        if index >= len(self.prospects):
            return None
        return self.prospects[index]

    def populate_form(self, prospect: Prospect) -> None:
        self.name_var.set(prospect.name)
        self.title_var.set(prospect.title)
        self.company_var.set(prospect.company)
        self.topics_var.set(", ".join(prospect.topics))
        self.profile_url_var.set(prospect.profile_url)
        self.status_var.set(prospect.status)
        self.note_var.set(prospect.note)

    def on_select(self, _: tk.Event) -> None:
        prospect = self.selected_prospect()
        if prospect is not None:
            self.populate_form(prospect)

    def clear_form(self) -> None:
        for variable in [self.name_var, self.title_var, self.company_var, self.topics_var, self.profile_url_var, self.status_var, self.note_var]:
            variable.set("")
        self.status_var.set("to_contact")
        self.listbox.selection_clear(0, tk.END)
        self.profile_text.delete("1.0", tk.END)
        self.write_results("")

    def current_prospect(self) -> Prospect:
        return Prospect(
            name=self.name_var.get().strip(),
            title=self.title_var.get().strip(),
            company=self.company_var.get().strip(),
            topics=normalize_topics(self.topics_var.get()),
            profile_url=self.profile_url_var.get().strip(),
            status=self.status_var.get().strip() or "to_contact",
            note=self.note_var.get().strip(),
            last_updated=now_utc(),
        )

    def save_current(self) -> None:
        prospect = self.current_prospect()
        if not prospect.name:
            messagebox.showerror("Missing name", "Please enter a prospect name.")
            return

        for index, existing in enumerate(self.prospects):
            if existing.name.lower() == prospect.name.lower():
                self.prospects[index] = prospect
                break
        else:
            self.prospects.append(prospect)

        save_store(self.store_path, self.prospects)
        self.reload()
        self.select_by_name(prospect.name)

    def delete_selected(self) -> None:
        prospect = self.selected_prospect()
        if prospect is None:
            messagebox.showinfo("Nothing selected", "Choose a prospect to delete.")
            return

        if not messagebox.askyesno("Delete prospect", f"Delete {prospect.name}?"):
            return

        self.prospects = [item for item in self.prospects if item.name.lower() != prospect.name.lower()]
        save_store(self.store_path, self.prospects)
        self.reload()
        self.clear_form()

    def select_by_name(self, name: str) -> None:
        for index, prospect in enumerate(self.prospects):
            if prospect.name.lower() == name.lower():
                self.listbox.selection_set(index)
                self.listbox.activate(index)
                self.listbox.see(index)
                self.populate_form(prospect)
                return

    def import_csv_dialog(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
        if not path:
            return
        self.prospects.extend(load_csv_rows(Path(path)))
        save_store(self.store_path, self.prospects)
        self.reload()

    def import_xlsx_dialog(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not path:
            return
        self.prospects.extend(load_xlsx_rows(Path(path)))
        save_store(self.store_path, self.prospects)
        self.reload()

    def export_xlsx_dialog(self) -> None:
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        export_xlsx_rows(Path(path), self.prospects)
        messagebox.showinfo("Export complete", f"Saved {len(self.prospects)} prospects to {path}")

    def write_results(self, text: str) -> None:
        self.result_box.configure(state="normal")
        self.result_box.delete("1.0", tk.END)
        self.result_box.insert("1.0", text)
        self.result_box.configure(state="disabled")

    def match_text_dialog(self) -> None:
        text = self.profile_text.get("1.0", tk.END).strip()
        if not text:
            messagebox.showinfo("Paste text", "Paste a profile, bio, or post into the text area first.")
            return

        keywords = extract_keywords(text)
        if not keywords:
            messagebox.showinfo("No keywords", "No meaningful keywords were found in the pasted text.")
            return

        ranked = rank_prospects(self.prospects, keywords)
        lines = []
        for score, prospect, matched in ranked[:10]:
            if score <= 0:
                continue
            matched_text = ", ".join(matched[:8]) if matched else "-"
            lines.append(f"{score:>3} | {prospect.name} | {prospect.title or '-'} | {prospect.company or '-'} | matched: {matched_text}")

        self.write_results("\n".join(lines) if lines else "No strong matches found.")


def launch_gui(args: argparse.Namespace) -> None:
    ProspectManagerApp(Path(args.store)).mainloop()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Manage a local prospect list for manual LinkedIn outreach.")
    parser.add_argument("--store", default=str(DEFAULT_STORE), help="Path to the local JSON file used to store prospects.")

    subparsers = parser.add_subparsers(dest="command", required=True)

    init_parser = subparsers.add_parser("init", help="Create an empty prospect store.")
    init_parser.set_defaults(func=init_store)

    gui_parser = subparsers.add_parser("gui", help="Open the desktop GUI.")
    gui_parser.set_defaults(func=launch_gui)

    add_parser = subparsers.add_parser("add", help="Add one prospect.")
    add_parser.add_argument("--name", required=True)
    add_parser.add_argument("--title", default="")
    add_parser.add_argument("--company", default="")
    add_parser.add_argument("--topics", default="")
    add_parser.add_argument("--profile-url", default="")
    add_parser.add_argument("--status", default="to_contact")
    add_parser.add_argument("--note", default="")
    add_parser.set_defaults(func=add_prospect)

    import_csv_parser = subparsers.add_parser("import-csv", help="Import prospects from a CSV file.")
    import_csv_parser.add_argument("csv")
    import_csv_parser.set_defaults(func=import_csv)

    import_xlsx_parser = subparsers.add_parser("import-xlsx", help="Import prospects from an Excel file.")
    import_xlsx_parser.add_argument("xlsx")
    import_xlsx_parser.set_defaults(func=import_xlsx)

    export_xlsx_parser = subparsers.add_parser("export-xlsx", help="Export prospects to an Excel file.")
    export_xlsx_parser.add_argument("xlsx")
    export_xlsx_parser.set_defaults(func=export_xlsx)

    list_parser = subparsers.add_parser("list", help="List all saved prospects.")
    list_parser.set_defaults(func=list_prospects)

    suggest_parser = subparsers.add_parser("suggest", help="Rank prospects by your interests.")
    suggest_parser.add_argument("--interests", required=True, help="Comma-separated topics you care about.")
    suggest_parser.add_argument("--limit", type=int, default=10)
    suggest_parser.set_defaults(func=suggest_prospects)

    match_parser = subparsers.add_parser("match-text", help="Rank prospects using pasted profile text.")
    match_parser.add_argument("--text", default=None, help="Profile or bio text. If omitted, reads from stdin.")
    match_parser.add_argument("--limit", type=int, default=10)
    match_parser.set_defaults(func=match_text)

    draft_parser = subparsers.add_parser("draft", help="Draft a personalized outreach message.")
    draft_parser.add_argument("--name", required=True)
    draft_parser.add_argument("--title", default="")
    draft_parser.add_argument("--company", default="")
    draft_parser.add_argument("--topics", default="")
    draft_parser.add_argument("--interests", default="")
    draft_parser.add_argument("--profile-url", default="")
    draft_parser.add_argument("--note", default="")
    draft_parser.add_argument("--your-name", required=True)
    draft_parser.set_defaults(func=draft_message)

    mark_parser = subparsers.add_parser("mark", help="Update a prospect's status.")
    mark_parser.add_argument("--name", required=True)
    mark_parser.add_argument("--status", required=True)
    mark_parser.add_argument("--note", default="")
    mark_parser.set_defaults(func=mark_status)

    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()